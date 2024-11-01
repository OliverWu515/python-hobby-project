import os
# import tkinter as tk
import csv
# import warnings
from openpyxl import load_workbook
import re
from tkinter import Tk, Label, Button, Entry, Frame, StringVar, scrolledtext, messagebox, filedialog

# regex patterns
re_bracket = re.compile(r'^\[(.+)]$')  # 检测中括号
re_number = re.compile(r'^\[(.+)]\[(.+)]$')  # 检测周数+地点(非实验课)/节数+周数(实验课)
re_lesson = re.compile(r'^第(\d+)-(\d+)节$')  # 检测节数
re_lesson_experiment = re.compile(r'^(\d+)-(\d+)节$')  # 检测节数(实验课)
re_name_experiment = re.compile(r'【实验】')  # 实验课程名称检测


# warnings.filterwarnings('error')
# functions


def select_file():
    selected_file_path = filedialog.askopenfilename(filetypes=[("xlsx表格", ".xlsx")])
    select_path.set(selected_file_path)


# core function
def process():
    temp = filepath_entry.get()
    temp = temp.replace('/', '\\')
    inputdir = os.path.dirname(temp)
    try:
        wb = load_workbook(temp)
    except Exception as e:
        messagebox.showerror(title='打开工作表时发生错误', message='错误信息：\n'
                                                         + str(e))
        return
    #     messagebox.showerror(title='提示', message='由于校网的Excel文档使用Apache POI制作，系统无法识别，\n'
    #                                              '需先用Excel/WPS打开并保存，才可继续操作。')
    #     return
    combine = False
    show_info_as_teacher = False
    if messagebox.askyesno("请选择", "是否将同一门实验的不同时段课程合并为一个名称?"):
        combine = True
    if combine and messagebox.askyesno("请选择", "是否将同一门实验的不同时段课程"
                                              "（如实验一/二等）放在”老师“一栏以方便查阅?"):
        show_info_as_teacher = True
    table = wb["sheet1"]
    # 获取表格行数
    nrows = table.max_row
    # 获取表格列数
    ncols = table.max_column

    ind = 0
    # 表头输入
    data = [["课程名称", "星期", "开始节数", "结束节数", "老师", "地点", "周数"]]

    for col_index in range(2, ncols + 1):
        for row_index in range(4, nrows + 1):
            if table.cell(row_index, col_index).value is not None:
                cell_values = table.cell(row_index, col_index).value
                cell_item = cell_values.split('\n')
                ind = data_process(cell_item, ind, col_index - 1, row_index - 3, data, combine, show_info_as_teacher)
                if ind == -1:
                    return
    if data[-1] is None:
        data[-1] = ['', '', '', '', '', '', '']
    transform_csv(data, os.path.join(inputdir, "res.csv"))
    insert_message("写入csv完成\n")


def data_process(inf, index, week, lesson_config, data, combine, show_info_as_teacher):
    for item in inf:
        try:
            if re.match(re_lesson, item):
                if data[index] is None:
                    continue
                data[index][2] = re.match(re_lesson, item).groups()[0]  # 开始节数
                data[index][3] = re.match(re_lesson, item).groups()[1]  # 结束节数
                if int(data[index][3]) - int(data[index][2]) >1 and int(data[index][2]) != lesson_config *2 - 1:
                    # 防重复项，对时间跨度超过2节的课程做特殊处理
                    data[index] = None
            elif re.match(re_number, item):
                if data[index] is None:
                    continue
                if re.match(re_number, item).groups()[0][-1] == '周':  # (非实验课)
                    data[index][6] = re.match(re_number, item).groups()[0][:-1]  # 周数
                    data[index][5] = re.match(re_number, item).groups()[1]  # 地点
                else:
                    data[index][6] = re.match(re_number, item).groups()[1][:-1]  # 周数
                    st_end = re.match(re_number, item).groups()[0]
                    data[index][2] = re.match(re_lesson_experiment, st_end).groups()[0]  # 开始节数
                    data[index][3] = re.match(re_lesson_experiment, st_end).groups()[1]  # 结束节数
                    if int(data[index][3]) - int(data[index][2]) >1 and int(data[index][2]) != lesson_config *2 - 1:
                        data[index] = None
            elif re.match(re_bracket, item):  # 有中括号,教师名/实验室名
                if data[index] is None:
                    continue
                if re.match(re_bracket, item).groups()[0][-1].isdigit():
                    data[index][5] = re.match(re_bracket, item).groups()[0]  # 实验室名
                else:
                    data[index][4] = re.match(re_bracket, item).groups()[0]  # 教师名
            else:  # 没有特殊格式,课程名
                if data[index] is None:
                    data[index] = ['', '', '', '', '', '', '']
                else:
                    index += 1
                    data.append(['', '', '', '', '', '', ''])
                item_modified = item
                if combine and re.match(re_name_experiment, item):  # 为实验课
                    item_modified = item.split(' ')[0]
                data[index][0] = item_modified  # 课程名
                data[index][1] = week  # 星期
                experiment_info = ""
                if show_info_as_teacher:
                    try:
                        experiment_info = "".join(item.split(' ')[1:])
                    except IndexError:
                        experiment_info = ""
                    finally:
                        data[index][4] = experiment_info
        except Exception as e:
            insert_message(str(e) + '\n发生错误时处理的项：' + item)
            return -1
    return index


# 转csv格式函数
def transform_csv(tab, outputdir):
    with open(outputdir, 'w', newline='', encoding='utf-8-sig') as f:
        write = csv.writer(f)
        data = []
        for i in tab:
            data.append(i)
        write.writerows(data)


def insert_message(message):
    scroll_inf.config(state="normal")
    scroll_inf.insert("end", message)
    scroll_inf.config(state="disabled")


def prompt():
    insert_message('使用方法：先从校网导出课表Excel文件，点击“选择文件”选择文件路径，'
                   '再点击“开始转换”转换后'
                   '会产生一个csv文件，将此csv文件导入至WakeUp课程表即可。\n\n'
                   "1.1版本更新说明：在原来的版本中同一门课程的实验由于课程名称中带有“实验一”等字样"
                   "而不会自动合并，导致“已添课程”中每节实验课都单独列出，不方便查看。现在每门课程的实验课在导出时会合并起来，方便查看。"
                   "同时通过将“实验一/二”等信息放在“老师”栏（实验课的“老师”栏往往空出），为不想要冗长的课程名称"
                   "又想要区分不同时段实验的同学提供了新选择。")


# GUI process
master_window = Tk()
master_window.title('HITsz课表信息提取程序for WakeUp课程表 V1.1(openpyxl version)')
master_window.geometry('480x320')
master_window.resizable(False, False)
select_path = StringVar()
Label(master_window, text='文件路径:').pack(fill="x")
filepath_entry = Entry(master_window, textvariable=select_path, state="disabled")
filepath_entry.pack(fill="x")
button_frame = Frame(master_window)
button_frame.pack(fill="x")
Button(button_frame, text='选择文件', command=select_file, width=15).pack(side="left", padx=20)
Button(button_frame, text='使用说明', command=prompt, width=15).pack(side="left", padx=25)
Button(button_frame, text='执行转换', command=lambda: process(), width=15).pack(side="right", padx=25)
scroll_inf = scrolledtext.ScrolledText(master_window, font=('楷体', 14))
scroll_inf.pack(side="bottom", pady=10, fill="x")
scroll_inf.config(state="disabled")
master_window.mainloop()

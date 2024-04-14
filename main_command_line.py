import os
import sys
import csv
import warnings
import openpyxl as xl
import re

# regex patterns
re_bracket = re.compile(r'^\[(.+)]$')  # 检测中括号
re_number = re.compile(r'^\[(.+)]\[(.+)]$')  # 检测周数+地点(非实验课)/节数+周数(实验课)
re_lesson = re.compile(r'^第(\d+)-(\d+)节$')  # 检测节数
re_lesson_experiment = re.compile(r'^(\d+)-(\d+)节$')  # 检测节数(实验课)
re_name_experiment = re.compile(r'【实验】')  # 实验课程名称检测

# functions
warnings.filterwarnings('error')


# core function
def process():
    prompt()
    temp = input("请输入文件路径：")
    temp = temp.replace('/', '\\')
    inputdir = os.path.dirname(temp)
    try:
        wb = xl.load_workbook(temp)
    except:
        print('提示: 由于校网的Excel文档使用Apache POI制作，系统无法识别，需先用Excel/WPS打开并保存，才可继续操作。系统将自行退出。')
        sys.exit(1)
    combine = False
    show_info_as_teacher = False
    if input("请选择: 是否将同一门实验的不同时段课程合并为一个名称?(y/n)") == 'y':
        combine = True
    if input("请选择: 是否将同一门实验的不同时段课程（如实验一/二等）放在”老师“一栏以方便查阅?(y/n)") == 'y':
        show_info_as_teacher = True
    table = wb["sheet1"]
    # 获取表格行数
    nrows = table.max_row
    # 获取表格列数
    ncols = table.max_column

    ind = 0
    data = [["课程名称", "星期", "开始节数", "结束节数", "老师", "地点", "周数"]]

    # 表头输入

    def data_process(inf, index, week):
        for item in inf:
            if re.match(re_lesson, item):
                data[index][2] = re.match(re_lesson, item).groups()[0]  # 开始节数
                data[index][3] = re.match(re_lesson, item).groups()[1]  # 结束节数
            elif re.match(re_number, item):
                if re.match(re_number, item).groups()[0][-1] == '周':  # (非实验课)
                    data[index][6] = re.match(re_number, item).groups()[0][:-1]  # 周数
                    data[index][5] = re.match(re_number, item).groups()[1]  # 地点
                else:
                    data[index][6] = re.match(re_number, item).groups()[1][:-1]  # 周数
                    st = re.match(re_number, item).groups()[0]
                    data[index][2] = re.match(re_lesson_experiment, st).groups()[0]  # 开始节数
                    data[index][3] = re.match(re_lesson_experiment, st).groups()[1]  # 结束节数
            elif re.match(re_bracket, item):  # 有中括号,教师名/实验室名
                if re.match(re_bracket, item).groups()[0][-1].isdigit():
                    data[index][5] = re.match(re_bracket, item).groups()[0]  # 实验室名
                else:
                    data[index][4] = re.match(re_bracket, item).groups()[0]  # 教师名
            else:  # 没有特殊格式,课程名
                index += 1
                data.append(['', '', '', '', '', '', ''])
                item_modified = item
                if combine and re.match(re_name_experiment, item):  # 为实验课
                    item_modified = item.split(' ')[0]
                data[index][0] = item_modified  # 课程名
                data[index][1] = week  # 星期
                if show_info_as_teacher:
                    try:
                        experiment_info = item.split(' ')[1]
                    except IndexError:
                        experiment_info = ""
                    finally:
                        data[index][4] = experiment_info
        return index

    for col_index in range(2, ncols + 1):
        for row_index in range(4, nrows + 1):
            if table.cell(row_index, col_index).value is not None:
                ar = table.cell(row_index, col_index).value
                lar = ar.split('\n')
                ind = data_process(lar, ind, col_index - 1)
    transform_csv(data, os.path.join(inputdir, "res.csv"))
    print("写入csv完成")


# 转csv格式函数
def transform_csv(tab, outputdir):
    with open(outputdir, 'w', newline='', encoding='utf-8-sig') as f:
        write = csv.writer(f)
        data = []
        for i in tab:
            data.append(i)
        write.writerows(data)


def prompt():
    print('使用方法：先从校网导出课表Excel文件，点击“选择文件”选择文件路径，'
          '再点击“开始转换”转换后'
          '会产生一个csv文件，将此csv文件导入至WakeUp课程表即可。\n'
          "1.1版本更新说明：在原来的版本中同一门课程的实验由于课程名称中带有“实验一”等字样"
          "而不会自动合并，导致“已添课程”中每节实验课都单独列出，不方便查看。现在每门课程的实验课在导出时会合并起来，方便查看。"
          "同时通过将“实验一/二”等信息放在“老师”栏（实验课的“老师”栏往往空出），为不想要冗长的课程名称"
          "又想要区分不同时段实验的同学提供了新选择。")


if __name__ == "__main__":
    process()
